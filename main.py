import class_set as cset
from function_set import *
import pandas as pd
import re
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Color, PatternFill, Font, Border, Side
from openpyxl import formatting, styles
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.formula import ArrayFormula

print('Reading data ...')

try:
    df= pd.read_excel('Input/Specific Examer.xlsx')
except:
    printErrorMsg('Can\'t find file \'Specific Examer.xlsx\'!')

MAIN_EXAMER_OF_ENG_SPEAKING = [x for x in df['English Speaking\n主考官'].tolist() if x == x]
ORAL_EXAMER_OF_ENG_SPEAKING = [x for x in df['English Speaking\nOral 考官'].tolist() if x == x]
MAIN_EXAMER_OF_ENG_LISTENING = [x for x in df['English Listening\n主考官'].tolist() if x == x]
MAIN_EXAMER_OF_CHIN_SPEAKING = [x for x in df['中文説話\n主考官'].tolist() if x == x]
ORAL_EXAMER_OF_CHIN_SPEAKING = [x for x in df['中文説話\nOral 考官'].tolist() if x == x]
MAIN_EXAMER_OF_CHIN_LISTENING = [x for x in df['中文聆聽\n主考官'].tolist() if x == x]
MAIN_EXAMER_OF_PTH = [x for x in df['普通話\n主考官'].tolist() if x == x]
MAIN_EXAMER_OF_VA = [x for x in df['VA\n主考官'].tolist() if x == x]
FOREIGN_TEACHER = [x for x in df['外籍老師'].tolist() if x == x]
SPECIAL_TIME_TEACHER = [x for x in df['特殊時數'].tolist() if x == x]
SPECIAL_TA = [x for x in df['特殊TA'].tolist() if x == x]
ENG_SPEAKING_HALL_TA = [x for x in df['English Speaking HALL TA'].tolist() if x == x]
SPEAKING_PR_TA = [x for x in df['English Speaking preparation room TA'].tolist() if x == x]

tmp = {}
for examer in MAIN_EXAMER_OF_VA:
    examer = examer.replace(' ', '')
    tmp[int(examer[:examer.index(':')])] = examer[examer.index(':')+1:]
MAIN_EXAMER_OF_VA = tmp

tmp = {}
for examer in SPECIAL_TIME_TEACHER:
    examer = examer.replace(' ', '')
    tmp[examer[:examer.index(':')]] = float(examer[examer.index(':')+1:])
SPECIAL_TIME_TEACHER = tmp

CANT_BE_EXAMER = [x for x in df['不能監考\n(校長)'].tolist() if x == x]

TA_DATA = []
for ta in [x for x in df['TA'].tolist() if x == x]:
    TA_DATA.append(cset.TA(name=ta))
    if ta in SPECIAL_TA:
        TA_DATA[-1].ratio = 0.75
        

try:
    df= pd.read_excel('Input/Other Info.xlsx', sheet_name='科目名稱對照')
except:
    printErrorMsg('Can\'t find file \'Other Info.xlsx\'!')

SUBJECT_NAME_DICT = {[x for x in df['科目中文名'].tolist() if x == x][i] : [x for x in df['科目縮寫'].tolist() if x == x][i].replace(' ','').split(',') for i in range(len([x for x in df['科目中文名'].tolist() if x == x]))}

df= pd.read_excel('Input/Other Info.xlsx', sheet_name='班別科室對照')
CLASS_DICT = {str([x for x in df['課室'].tolist() if x == x][i]) : [x for x in df['班別'].tolist() if x == x][i] for i in range(len([x for x in df['課室'].tolist() if x == x]))}

df= pd.read_excel('Input/Other Info.xlsx', sheet_name='班主任')
CLASS_TEACHER = {[x for x in df['班主任'].tolist() if x == x][i] : [x for x in df['班別'].tolist() if x == x][i] for i in range(len([x for x in df['班主任'].tolist() if x == x]))}

try:
    df= pd.read_excel('Input/Exam Timetable.xlsx', skiprows=[0], usecols=lambda x: 'Unnamed' not in x)
except:
    printErrorMsg('Can\'t find file \'Exam Timetable.xlsx\'!')

ET_DATA = []
for date in df.columns:
    tmp = date.replace(' ','')
    tmp = tmp.replace('（','(')
    tmp = tmp.replace('）',')')
    df.rename(columns={date: tmp}, inplace=True)
    df.replace(date, tmp, inplace=True)
    date = tmp
    if date[-2:-1] in ['一', '二', '三', '四', '五', '六', '日']:
        ET_DATA.append(cset.exam(examDate=date))
    else:
        print('Exam Timetable: Date Formate Error!')

for exam in ET_DATA:
    exam.subjects = []
    exam.noExam = []
    listedColum = df[exam.examDate].tolist()
    listedColum.insert(0, exam.examDate)
    form = 0
    for i in range(len(listedColum)-1):
        if listedColum[i] == exam.examDate:
            form += 1
            if listedColum[i+1] == '上課':
                exam.noExam.append(form)
    listedColum = [x for x in listedColum if x == x]
    listedColum = list(filter(lambda i: i != '上課', listedColum))
    i = 0
    form = 0
    while i < len(listedColum):
        if listedColum[i] == exam.examDate:
            form += 1
            i += 1
        else:
            exam.subjects.append(cset.subject(name = listedColum[i], timeLimit = listedColum[i+1], room = listedColum[i+2], period = transferTimeFormat(listedColum[i+3]), form=form, parent=exam))
            i += 4

def appendSubjectsAndClasses(teacher, lessonName):
    s = re.compile(r'[1-6][A-F] [A-Za-z0-9]+').findall(lessonName)
    c = re.compile(r'[1-6][A-F]').findall(lessonName)
    if len(s) > 0:
        if s[0][3:] not in [key for key in teacher.teachedSubjectsAndClasses]:
            teacher.teachedSubjectsAndClasses[s[0][3:]] = []
        if not set(c).issubset(teacher.teachedSubjectsAndClasses[s[0][3:]]):
            teacher.teachedSubjectsAndClasses[s[0][3:]] += c
        return c, s[0][3:]
    else:
        return None

try:
    sheets = pd.ExcelFile('Input/Teacher Timetable.xlsx')
except:
    printErrorMsg('Can\'t find file \'Teacher Timetable.xlsx\'!')


TT_DATA = []
dateDict = {'Mon' : '一', 'Tue' : '二', 'Wed' : '三', 'Thu' : '四', 'Fri' : '五'}
timeSlot = pd.read_excel('Input/Teacher Timetable.xlsx', skiprows=[0,1])['Unnamed: 0'].tolist()
timeSlot = list(map(transferTimeFormat, timeSlot))

for sheetName in sheets.sheet_names:
    if sheetName not in CANT_BE_EXAMER:
        TT_DATA.append(cset.teacher(sheetName))
for teacher in TT_DATA:
    teacher.ratio = 1
    teacher.lessons = {}
    teacher.totalTime = 0
    teacher.teachedSubjectsAndClasses = {}
    teacher.exams = {}
    df = pd.read_excel('Input/Teacher Timetable.xlsx', skiprows=[0,1], sheet_name=teacher.name, usecols=lambda x: 'Unnamed' not in x)

    for date in df.columns:
        listedColum = df[date].tolist()
        teacher.lessons[dateDict[date]] = []
        for i in range(len(listedColum)):
            if 'unch' in str(listedColum[i]):
                break
            elif 'orning' in str(listedColum[i]) and teacher.name in [key for key in CLASS_TEACHER]:
                teacher.lessons[dateDict[date]].append(cset.lesson(name='班主任', period=timeSlot[i], classes=[CLASS_TEACHER[teacher.name]], room=[i for i in CLASS_DICT if CLASS_DICT[i] == CLASS_TEACHER[teacher.name]]))
            elif re.match('[1-6][A-F]', str(listedColum[i])) != None:
                classes, name = appendSubjectsAndClasses(teacher, listedColum[i])
                room = listedColum[i].split(' ')[-1]
                teacher.lessons[dateDict[date]].append(cset.lesson(name=name, period=timeSlot[i], classes=classes, room=room))
    
    if teacher.name in [key for key in SPECIAL_TIME_TEACHER]:
        teacher.ratio = SPECIAL_TIME_TEACHER[teacher.name]

for teacher in TT_DATA:
    teacher.totalTime = 0
    for exam in ET_DATA:
        for needLessonForms in exam.noExam:
            for lesson in teacher.lessons[exam.examDate[-2:-1]]:
                if lesson.classes[0][0] == str(needLessonForms) and lesson.name != '班主任':
                    teacher.lessonTime += 35
        teacher.exams[exam.examDate] = []
    teacher.totalTime = teacher.lessonTime

for ta in TA_DATA:
    for exam in ET_DATA:
        ta.exams[exam.examDate] = []

AVG_TIME = 0

def findAvalibleTeachers(subject, specificExamer=None, needCheck = False):
    avalibleTeachersList = []
    teacherData = []
    if specificExamer != None:
        for teacherNames in specificExamer:
            teacherData.append(findParentObj(TT_DATA, teacherNames))
    else:
        teacherData = sorted(TT_DATA, key=lambda x: x.ratio, reverse=False)
    for teacher in teacherData:
        avalible = True
        if teacher.name in FOREIGN_TEACHER and specificExamer != FOREIGN_TEACHER:
            avalible = False
        if specificExamer == None and teacher.name in [key for key in SPECIAL_TIME_TEACHER] and (teacher.totalTime + subject.timeLimit[0] - 20) >= AVG_TIME*teacher.ratio:
            avalible = False
        if len(subject.parent.noExam) > 0:
            for lesson in filter(lambda x: int(x.classes[0][0]) in subject.parent.noExam, teacher.lessons[subject.parent.examDate[-2:-1]]):
                if avalible:
                    avalible = checkTime(subject.period, lesson.period)
                else:
                    break
        if subject.parent.examDate in [key for key in teacher.exams]:
            for examTime in map(lambda x: x.period, teacher.exams[subject.parent.examDate]):
                if avalible:
                    avalible = checkTime(subject.period, examTime)
                else:
                    break
        
        if needCheck:
            tmp = subject.name[:subject.name.index(' ') if ' ' in subject.name else len(subject.name)]
            if tmp in [key for key in SUBJECT_NAME_DICT]:
                for subjectName in SUBJECT_NAME_DICT[tmp]:
                    if subjectName in [key for key in teacher.teachedSubjectsAndClasses]:
                        if len(subject.room) > 3:
                            if CLASS_DICT[subject.room[subject.teachers.index('')]] in teacher.teachedSubjectsAndClasses[subjectName]:
                                avalible = False
                        else:
                            if str(subject.form) in list(map(lambda x: x[0], teacher.teachedSubjectsAndClasses[subjectName])):
                                avalible = False

        if avalible:
            avalibleTeachersList.append(teacher)
            
    avalibleTeachersList.sort(key=lambda x: x.totalTime, reverse=False)
    return avalibleTeachersList[0]

def appendTeachers(i, subject, avalibleTeacher, isOral=False):
    if subject.teachers[i] != '':
        return
    subject.teachers[i] = avalibleTeacher.name
    tmp = 0 if not isOral else 1
    avalibleTeacher.totalTime += subject.timeLimit[tmp]
    avalibleTeacher.exams[subject.parent.examDate].append(cset.examDetails(subject.name, subject.period.split('\n')[0 if not isOral else 1], subject.room[i], subject.timeLimit[tmp]))

def appendTA(i, subject, specific=None):
    if specific == []:
        return
    avalibleTAList = []
    tmp = []
    if specific != None:
        for TAName in specific:
            tmp.append(findParentObj(TA_DATA, TAName))
    else:
        tmp = TA_DATA

    for TA in tmp:
        avalible = True
        if subject.parent.examDate in [key for key in TA.exams]:
            for examTime in map(lambda x: x.period, TA.exams[subject.parent.examDate]):
                if avalible:
                    avalible = checkTime(subject.period, examTime)
                else:
                    break
        
        if specific == None and TA.name in SPECIAL_TA:
            avalible = False 

        if avalible:
            avalibleTAList.append(TA)

    if len(avalibleTAList) == 0:
        for TAName in SPECIAL_TA:
            avalibleTAList.append(findParentObj(TA_DATA, TAName))

    avalibleTAList.sort(key=lambda x: x.totalTime, reverse=False)
    avalibleTA = avalibleTAList[0]
    subject.teachers[i] = avalibleTA.name
    avalibleTA.totalTime += subject.timeLimit[0]
    avalibleTA.exams[subject.parent.examDate].append(cset.examDetails(subject.name, subject.period.split('\n')[0], subject.room[i], subject.timeLimit[0]))

print('Processing ...')

for exam in ET_DATA:
    for subject in exam.subjects:
        if 'peaking' in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, MAIN_EXAMER_OF_ENG_SPEAKING))
            AVG_TIME += subject.timeLimit[0]
            for i in range(1,len(list(filter(lambda x: x == 'HALL', subject.room)))):
                appendTA(i, subject, ENG_SPEAKING_HALL_TA)
            for i in range(subject.teachers.index(''),len(subject.room)):
                if 'p' in subject.room[i]:
                    appendTA(i, subject, SPEAKING_PR_TA)
            for i in range(len(FOREIGN_TEACHER)):
                try:
                    appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject, FOREIGN_TEACHER), isOral=True)
                    AVG_TIME += subject.timeLimit[1]
                except:
                    print('FOREIGN_TEACHER not avalible')
            for i in range(len(list(filter(lambda x: x == '', subject.teachers)))):
                appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject, ORAL_EXAMER_OF_ENG_SPEAKING), isOral=True)
                AVG_TIME += subject.timeLimit[1]
        elif '說話' in subject.name or '説話' in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, MAIN_EXAMER_OF_CHIN_SPEAKING))
            AVG_TIME += subject.timeLimit[0]
            for i in range(1,len(list(filter(lambda x: x == 'HALL', subject.room)))):
                appendTA(i, subject)
            for i in range(subject.teachers.index(''),len(subject.room)):
                if subject.room[i][-2] == 'p':
                    appendTA(i, subject, SPEAKING_PR_TA)
            for i in range(len(list(filter(lambda x: x == '', subject.teachers)))):
                appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject, ORAL_EXAMER_OF_CHIN_SPEAKING), isOral=True)
                AVG_TIME += subject.timeLimit[1]
            
for exam in ET_DATA:
    for subject in exam.subjects:
        if '普通話' in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, MAIN_EXAMER_OF_PTH))
            appendTA(len(subject.room)-1, subject)
            AVG_TIME += subject.timeLimit[0]
            for i in range(len(list(filter(lambda x: x == '', subject.teachers)))):
                appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject))
                AVG_TIME += subject.timeLimit[0]
        elif '聆聽' in subject.name and 'TSA' not in subject.name and '普通話' not in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, MAIN_EXAMER_OF_CHIN_LISTENING))
            appendTA(len(subject.room)-1, subject)
            AVG_TIME += subject.timeLimit[0]
        elif 'istening' in subject.name and 'TSA' not in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, MAIN_EXAMER_OF_ENG_LISTENING))
            AVG_TIME += subject.timeLimit[0]
            appendTA(2, subject)
        elif '視覺藝術' in subject.name:
            appendTeachers(0, subject, findAvalibleTeachers(subject, [MAIN_EXAMER_OF_VA[subject.form]]))
            AVG_TIME += subject.timeLimit[0]
        elif 'HALL' in subject.room:
            appendTA(len(subject.room)-1, subject)            

for exam in ET_DATA:
    for subject in exam.subjects:
        AVG_TIME += subject.timeLimit[0] * len(list(filter(lambda x: x == '', subject.teachers)))
        
for teacher in TT_DATA:
    AVG_TIME += teacher.lessonTime * teacher.ratio


AVG_TIME /= (sum(list(map(lambda x: x.ratio, TT_DATA+TA_DATA))))

for subject in sorted(list(filter(lambda x: '' in x.teachers, list(np.concatenate(list(map(lambda x: x.subjects, ET_DATA))).flat))), key=lambda x: x.timeLimit[0], reverse=True):
    if 'HALL' in subject.room:
        appendTeachers(0, subject, findAvalibleTeachers(subject))   
        if '' in subject.teachers:
            appendTA(subject.teachers.index(''), subject, specific=list(filter(lambda x: (findParentObj(TA_DATA, x).totalTime + subject.timeLimit[0] - 20) < AVG_TIME*findParentObj(TA_DATA, x).ratio, SPECIAL_TA)))
        for i in range(len(list(filter(lambda x: x == '', subject.teachers)))):
            appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject))
    else:
        for i in range(len(list(filter(lambda x: x == '', subject.teachers)))):
            appendTeachers(subject.teachers.index(''), subject, findAvalibleTeachers(subject, needCheck=True))

offset = 20

print('Outputting ...')

workbook = openpyxl.Workbook()
sheet = workbook.worksheets[0]
sheet.title = '考試時間表 + 監考'

formDict = { 1 : '中一級', 2 : '中二級', 3 : '中三級', 4 : '中四級', 5 : '中五級', 6 : '中六級'}

greyFill = PatternFill(patternType='solid', fgColor=Color(rgb='D9D9D9'))
yellowFill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFF00'))
orangeFill = PatternFill(patternType='solid', fgColor=Color(rgb='FFC000'))
mediumBorder = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
thinBorder = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
bottemBorder = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style=None), 
                     bottom=Side(style='medium'))

red_font = styles.Font(size=14, bold=True, color='9c0103')
red_fill = styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_font = styles.Font(size=14, bold=True, color='006100')
green_fill = styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')


for i in range(ET_DATA[0].subjects[-1].form):
    sheet.cell(row = sheet.max_row+2, column = 1).value = formDict[i+1]
    sheet.cell(row = sheet.max_row+1, column = 1).border = thinBorder
    top = sheet.max_row

    subjectListFilteredByForm = list(map(lambda x: list(filter(lambda y: y.form == i+1, x)) ,map(lambda x: x.subjects, ET_DATA)))
    maxRowLength = [''] * len(sorted(subjectListFilteredByForm, key=lambda z: len(z), reverse=True)[0])
    
    for j in range(len(maxRowLength)):
        maxRowLength[j] = max(list(map(lambda x: len(x[j].room) if j < len(x) else 0, subjectListFilteredByForm)))
    for j in range(len(maxRowLength)):
        if j == 0:
            tmp = sheet.max_row
        else:
            tmp = sheet.max_row+1
        for col in range(len(ET_DATA)+1):
            for row in range(maxRowLength[j]+3):
                sheet.cell(row = tmp+row+1, column = col+1).border = thinBorder
                sheet.cell(row = tmp+row+1, column = col+1).alignment = Alignment(horizontal='center', wrapText=True, vertical = 'center')
                
    for col, exam in enumerate(ET_DATA,start=2):
        sheet.cell(row = top, column = col).value = exam.examDate
        sheet.cell(row = top, column = col).border = thinBorder
        sheet.cell(row = top, column = col).alignment = Alignment(horizontal='center', wrapText=True, vertical = 'center')
        sheet.cell(row = top, column = col).font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(col)].width = 17
        current_row = top+1
        for k, subject in enumerate(list(filter(lambda x: x.form == i+1, exam.subjects))):
            sheet.cell(row = current_row, column = col).value = subject.name
            sheet.cell(row = current_row, column = col).font = Font(bold=True)
            sheet.cell(row = current_row, column = 1).value = '科 目'
            sheet.cell(row = current_row, column = 1).fill = orangeFill
            sheet.cell(row = current_row, column = col).fill = orangeFill

            sheet.cell(row = current_row+1, column = col).value = '/'.join(list(map(lambda x: str(x) ,subject.timeLimit)))
            sheet.cell(row = current_row+1, column = 1).value = '時 限'

            sheet.cell(row = current_row+2, column = col).value = subject.period
            sheet.cell(row = current_row+2, column = 1).value = '應考時間'

            current_row += 3
            if sheet.cell(row = current_row, column = 1).value == None:
                sheet.cell(row = current_row, column = 1).value = '應考試場\n監考'
                sheet.cell(row = current_row, column = 1).fill = yellowFill
                sheet.cell(row = current_row, column = 1).border = thinBorder
                sheet.cell(row = current_row, column = 1).alignment = Alignment(horizontal='center', wrapText=True, vertical = 'center')
            for j in range(maxRowLength[k]+1):
                if j < len(subject.room):
                    sheet.cell(row = current_row, column = col).value = subject.room[j] + ': ' + subject.teachers[j]
                    sheet.cell(row = current_row, column = col).fill = yellowFill
                    sheet.cell(row = current_row, column = col).alignment = Alignment(horizontal='center', wrapText=True, vertical = 'center')
                    sheet.cell(row = current_row, column = col).border = thinBorder
                else:
                    sheet.cell(row = current_row, column = col).value = None
                current_row += 1
            if len(subject.room) == maxRowLength[k]:
                sheet.merge_cells(start_row=(current_row-1-maxRowLength[k]), start_column=1, end_row=current_row-2, end_column=1)
    for y in range(2, sheet.max_column+1):
        for x in range(top, sheet.max_row+1):
            if sheet.cell(row = x, column = y).value == None and sheet.cell(row = x, column = y).border == thinBorder:
                sheet.cell(row = x, column = y).fill = greyFill

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------

workbook.create_sheet('老師上課 + 監考時數')
sheet2 = workbook.worksheets[1]
sheet2.sheet_view.zoomScale = 70

header = ['老師', '班主任', '比例', '上課時數', '考試時數', '總時數', '平均', '誤差']
for i, lable in enumerate(header, start=1):
    sheet2.cell(row = 1, column = i).value = lable

for i, examDate in enumerate(map(lambda x: x.examDate, ET_DATA)):
    sheet2.cell(row = 1, column = sheet2.max_column+1).value = examDate
    sheet2.column_dimensions[get_column_letter(sheet2.max_column)].width = 18
    sheet2.column_dimensions[get_column_letter(sheet2.max_column+1)].width = 18
    sheet2.column_dimensions[get_column_letter(sheet2.max_column+2)].width = 10
    sheet2.column_dimensions[get_column_letter(sheet2.max_column+3)].width = 18
    sheet2.column_dimensions[get_column_letter(sheet2.max_column+4)].width = 6
    sheet2.merge_cells(start_row=1, start_column=sheet2.max_column, end_row=1, end_column=sheet2.max_column+4)
    
for y in range(1, sheet2.max_column+1):
        sheet2.cell(row = sheet2.max_row, column = y).border = bottemBorder

TA_tmp = []

for teacher in TT_DATA+TA_DATA:
    current_col = 1
    sheet2.cell(row = sheet2.max_row+1, column = current_col).value = teacher.name
    if teacher.name in [key for key in CLASS_TEACHER]:
        sheet2.cell(row = sheet2.max_row, column = current_col+1).value = CLASS_TEACHER[teacher.name]
    sheet2.cell(row = sheet2.max_row, column = current_col+2).value = teacher.ratio
    if teacher.ratio != 1:
        sheet2.cell(row = sheet2.max_row, column = current_col+2).fill = yellowFill
    sheet2.cell(row = sheet2.max_row, column = current_col+3).value = teacher.lessonTime if type(teacher) == cset.teacher else 0
    sheet2.cell(row = sheet2.max_row, column = current_col+4).value = teacher.totalTime - (teacher.lessonTime if type(teacher) == cset.teacher else 0)
    sheet2.cell(row = sheet2.max_row, column = current_col+5).value = teacher.totalTime
    
    if type(teacher) == cset.TA and teacher.ratio == 0:
        TA_tmp.append(sheet2.cell(row = sheet2.max_row, column = current_col+5).coordinate)
        
    # sheet2.cell(row = sheet2.max_row, column = current_col+6).value = round(AVG_TIME * teacher.ratio)
    sheet2.cell(row = sheet2.max_row, column = current_col+7).value = '={}-{}'.format(sheet2.cell(row = sheet2.max_row, column = current_col+5).coordinate, sheet2.cell(row = sheet2.max_row, column = current_col+6).coordinate)
    sheet2.conditional_formatting.add(sheet2.cell(row = sheet2.max_row, column = current_col+7).coordinate, formatting.rule.CellIsRule(operator='notBetween', formula=[str(-offset),str(offset)], fill=red_fill, font=red_font))
    sheet2.conditional_formatting.add(sheet2.cell(row = sheet2.max_row, column = current_col+7).coordinate, formatting.rule.CellIsRule(operator='between', formula=[str(-offset),str(offset)], fill=green_fill, font=green_font))
    sheet2.row_dimensions[sheet2.max_row].height = 32
    current_col = 9
    tmp = sheet2.max_row
    for exam in ET_DATA:
        current_row = tmp
        if len(exam.noExam) > 0 and type(teacher) == cset.teacher:
            for lesson in filter(lambda x: int(x.classes[0][0]) in exam.noExam, teacher.lessons[exam.examDate[-2:-1]]):
                sheet2.cell(row = current_row, column = current_col).value = lesson.period
                sheet2.cell(row = current_row, column = current_col+1).value = lesson.name
                sheet2.cell(row = current_row, column = current_col+2).value = ','.join(lesson.classes)
                sheet2.cell(row = current_row, column = current_col+3).value = ''.join(lesson.room)
                if lesson.name != '班主任':
                    sheet2.cell(row = current_row, column = current_col+4).value = 35
                
                for col in range(current_col, current_col+5):
                    sheet2.cell(row = current_row, column = col).font = Font(color='0066FF', bold=True)
                
                sheet2.row_dimensions[current_row].height = 32

                current_row += 1
        for examDetails in teacher.exams[exam.examDate]:
            sheet2.cell(row = current_row, column = current_col).value = examDetails.period
            sheet2.cell(row = current_row, column = current_col+1).value = examDetails.name
            # sheet2.cell(row = current_row, column = current_col+2).value = 
            sheet2.cell(row = current_row, column = current_col+3).value = ''.join(examDetails.room)
            sheet2.cell(row = current_row, column = current_col+4).value = examDetails.timeLimit
            
            sheet2.row_dimensions[current_row].height = 32

            current_row += 1

        current_col += 5

    for i in range(len(header)):
        sheet2.merge_cells(start_row=tmp, end_row=sheet2.max_row, start_column=i+1, end_column=i+1)

    for y in range(1, sheet2.max_column+1):
        sheet2.cell(row = sheet2.max_row, column = y).border = bottemBorder

for i in range(len(ET_DATA)+1):
    for x in range(1, sheet2.max_row+1):
        sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border = Border(left=Side(style=None) if sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.left.style == None else sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.left, 
                     right=Side(style='medium'), 
                     top=Side(style=None) if sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.top.style == None else sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.top, 
                     bottom=Side(style=None) if sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.bottom.style == None else sheet2.cell(row = x, column = (i*5+8) if i != 0 else len(header)).border.bottom)

for y in range(1, sheet2.max_column+1):
    for x in range(1, sheet2.max_row+1):
        sheet2.cell(row = x, column = y).alignment = Alignment(horizontal='center', wrapText=True, vertical = 'center')
        sheet2.cell(row = x, column = y).font = Font(size=12, name='Times New Roman', color=sheet2.cell(row = x, column = y).font.color, bold=sheet2.cell(row = x, column = y).font.bold)
        sheet2.cell(row = x, column = y).border = Border(left=Side(style='thin') if sheet2.cell(row = x, column = y).border.left.style == None else sheet2.cell(row = x, column = y).border.left, 
                     right=Side(style='thin') if sheet2.cell(row = x, column = y).border.right.style == None else sheet2.cell(row = x, column = y).border.right, 
                     top=Side(style='thin') if sheet2.cell(row = x, column = y).border.top.style == None else sheet2.cell(row = x, column = y).border.top, 
                     bottom=Side(style='thin') if sheet2.cell(row = x, column = y).border.bottom.style == None else sheet2.cell(row = x, column = y).border.bottom)
        
avg_formula = '=ROUND((SUM($D$2:$D${}*$C$2:$C${})'.format(str(sheet2.max_row), str(sheet2.max_row))
for coor in TA_tmp:
    avg_formula += '-'+coor
avg_formula += '+SUM($E$2:$E${}))/SUM($C$2:$C${}),0)'.format(str(sheet2.max_row), str(sheet2.max_row))
tmp = sheet2.cell(row = sheet2.max_row+1, column=7).coordinate
sheet2[tmp] = ArrayFormula(tmp, avg_formula)

for y in range(3, 7):
    tmp = sheet2.cell(row = sheet2.max_row, column=y).coordinate
    sheet2[tmp] = ArrayFormula(tmp, '=SUM({}:{})'.format(sheet2.cell(row = 2, column=y).coordinate, sheet2.cell(row = sheet2.max_row-1, column=y).coordinate))

for x in range(2, sheet2.max_row):
    if not isinstance(sheet2.cell(row = x, column = 7), MergedCell):
        sheet2.cell(row = x, column = 7).value = '=ROUND({}*{}, 0)'.format(str(sheet2.cell(row = sheet2.max_row, column = 7).coordinate), str(sheet2.cell(row = x, column = 3).coordinate))

sheet2.freeze_panes = sheet2.cell(row = 2, column = len(header)+1).coordinate
workbook.save('監考時間表.xlsx')

